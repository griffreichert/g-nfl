"""Parse the Cville pool standings workbook into tidy pick rows.

The workbook has one sheet per week ("Week 1".."Week 18" plus playoff
rounds). Each week sheet contains, anchored to fixed rows:

- row 5: pick-slot headers (BB, 2..6, U-DOG, SD, MON — playoff rounds
  use their own headers: 1..6, Gm 1/Total, Spread/Total)
- rows 6+: one row per picker, col A = picker name, then one team
  abbreviation per slot ("--" or blank = no pick)
- a per-slot W/L result grid further right (col Z+), headers mirroring
  the pick-slot headers
- a lines block lower down (marker cell "Away" in col B): one row per
  game `away | line | home`, where the line is the AWAY team's spread
  (KC -3.0 LAC = road favorite KC by 3)

Every team is run through `standardize_teams`, and each pick gets a
nflverse-style game_id (`2025_01_KC_LAC`) derived from the lines block,
so pool picks join cleanly to schedules/results.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from g_nfl.utils.teams import standardize_teams

HEADER_ROW = 5
FIRST_PICKER_ROW = 6
MAX_SLOT_COL = 12  # pick slots live in cols B..L
RESULT_PICKER_COL = 26  # per-slot W/L grid: picker names in col Z
NO_PICK = {"", "--", "-", "NONE", "NO PICK"}

WEEK_SHEET_RE = re.compile(r"^Week (\d+)$")
# nflverse week numbers for the 2021+ 18-game era
PLAYOFF_WEEKS = {
    "Wild Card": 19,
    "Divisional": 20,
    "Conference": 21,
    "SuperBowl": 22,
}

# header cell -> (slot, pick_type); numbered headers handled separately
SLOT_HEADERS = {
    "BB": ("bb", "best_bet"),
    "U-DOG": ("udog", "underdog"),
    "SD": ("sd", "survivor"),
    "MON": ("mnf", "mnf"),
    "MNF": ("mnf", "mnf"),
    "SPREAD": ("spread", "regular"),
    "GM 1": ("gm1", "regular"),
    "GM 2": ("gm2", "regular"),
}


def _clean(value: Any) -> str:
    """Normalize a cell to a stripped upper-case string ('' if empty)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip().upper()


def _slot_columns(ws) -> list[tuple[int, str, str]]:
    """Read the header row into [(col, slot, pick_type), ...].

    Numbered headers (1..6) are regular ATS picks; a 'Total' header is an
    over/under on the game of the preceding slot (Conference, SuperBowl).
    """
    slots: list[tuple[int, str, str]] = []
    for col in range(2, MAX_SLOT_COL + 1):
        header = _clean(ws.cell(row=HEADER_ROW, column=col).value)
        if not header:
            continue
        if header in SLOT_HEADERS:
            slot, pick_type = SLOT_HEADERS[header]
        elif header.isdigit():
            slot, pick_type = header, "regular"
        elif header == "TOTAL":
            prev = slots[-1][1] if slots else "game"
            slot, pick_type = f"{prev}_total", "total"
        else:
            raise ValueError(f"unrecognized pick header {header!r} in col {col}")
        slots.append((col, slot, pick_type))
    return slots


def _picker_rows(ws) -> list[tuple[int, str]]:
    """Rows of the pick grid: contiguous non-empty col-A cells from row 6."""
    rows = []
    for row in range(FIRST_PICKER_ROW, ws.max_row + 1):
        picker = str(ws.cell(row=row, column=1).value or "").strip()
        if not picker:
            break
        rows.append((row, picker))
    return rows


def _games(ws, season: int, week: int) -> list[dict]:
    """Parse the lines block: [{away, home, line, game_id}, ...].

    The line is quoted for the AWAY team (negative = road favorite).
    """
    marker_row = None
    for row in range(FIRST_PICKER_ROW, ws.max_row + 1):
        if _clean(ws.cell(row=row, column=2).value) == "AWAY":
            marker_row = row
            break
    if marker_row is None:
        return []

    games = []
    for row in range(marker_row + 1, ws.max_row + 1):
        away_raw = _clean(ws.cell(row=row, column=2).value)
        home_raw = _clean(ws.cell(row=row, column=4).value)
        try:
            # some sheets store lines as text ('-3.5')
            line = float(ws.cell(row=row, column=3).value)
        except (TypeError, ValueError):
            continue
        if away_raw == "OVER":
            # over/under row: the total line for the game above it
            if games:
                games[-1]["total"] = line
            continue
        away = standardize_teams(away_raw)
        home = standardize_teams(home_raw)
        games.append(
            {
                "away": away,
                "home": home,
                "line": line,
                "total": None,
                "game_id": f"{season}_{week:02d}_{away}_{home}",
            }
        )
    return games


def _result_columns(ws, slots: list[tuple[int, str, str]]) -> dict[str, int]:
    """Map slot -> column of the per-slot W/L grid, if its headers match."""
    headers = {}
    for col, slot, _ in slots:
        pick_header = _clean(ws.cell(row=HEADER_ROW, column=col).value)
        result_col = None
        for rcol in range(RESULT_PICKER_COL + 1, RESULT_PICKER_COL + 15):
            if (
                _clean(ws.cell(row=HEADER_ROW, column=rcol).value) == pick_header
                and rcol not in headers.values()
            ):
                result_col = rcol
                break
        if result_col is not None:
            headers[slot] = result_col
    return headers


def parse_sheet(ws, season: int, week: int, week_label: str) -> list[dict]:
    """Parse one week sheet into pick rows."""
    slots = _slot_columns(ws)
    games = _games(ws, season, week)
    by_team = {g["away"]: g for g in games} | {g["home"]: g for g in games}
    result_cols = _result_columns(ws, slots)
    # totals attach to a game by position: gm1 -> 1st line row, gm2 -> 2nd
    total_game = {"gm1_total": 0, "gm2_total": 1, "spread_total": 0}

    picks = []
    for row, picker in _picker_rows(ws):
        for col, slot, pick_type in slots:
            value = _clean(ws.cell(row=row, column=col).value)
            if value in NO_PICK:
                continue

            game, team, spread = None, None, None
            if pick_type == "total":
                if value not in ("OVER", "UNDER"):
                    raise ValueError(
                        f"{week_label} {picker} {slot}: expected OVER/UNDER, got {value!r}"
                    )
                team = value
                idx = total_game.get(slot, 0)
                game = games[idx] if idx < len(games) else None
                # for totals, "spread" carries the over/under line
                spread = game["total"] if game else None
            else:
                team = standardize_teams(value)
                game = by_team.get(team)
                if game is not None:
                    spread = game["line"] if team == game["away"] else -game["line"]

            result = None
            if slot in result_cols:
                cell = _clean(ws.cell(row=row, column=result_cols[slot]).value)
                if cell in ("W", "L", "T", "P"):
                    result = cell

            picks.append(
                {
                    "season": season,
                    "week": week,
                    "week_label": week_label,
                    "picker": picker,
                    "slot": slot,
                    "pick_type": pick_type,
                    "team_picked": team,
                    "spread": spread,
                    "game_id": game["game_id"] if game else None,
                    "result": result,
                }
            )
    return picks


def parse_workbook(path: Path | str, season: int) -> list[dict]:
    """Parse every week sheet in the standings workbook into pick rows."""
    wb = openpyxl.load_workbook(path, data_only=True)
    picks = []
    for name in wb.sheetnames:
        match = WEEK_SHEET_RE.match(name.strip())
        if match:
            week = int(match.group(1))
        elif name.strip() in PLAYOFF_WEEKS:
            week = PLAYOFF_WEEKS[name.strip()]
        else:
            continue
        picks.extend(parse_sheet(wb[name], season, week, name.strip()))
    return picks


def main() -> None:
    """Demo: parse a workbook and print summary counts."""
    import argparse
    from collections import Counter

    ap = argparse.ArgumentParser(description="Parse pool standings workbook")
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--season", type=int, required=True)
    args = ap.parse_args()

    picks = parse_workbook(args.xlsx, args.season)
    print(f"{len(picks)} picks parsed")
    print("\nby week:", dict(sorted(Counter(p["week"] for p in picks).items())))
    print("\nby type:", dict(Counter(p["pick_type"] for p in picks)))
    print("\npickers:", sorted({p["picker"] for p in picks}))
    graded = sum(1 for p in picks if p["result"])
    print(
        f"\nwith result: {graded}, with game_id: "
        f"{sum(1 for p in picks if p['game_id'])}"
    )


if __name__ == "__main__":
    main()
