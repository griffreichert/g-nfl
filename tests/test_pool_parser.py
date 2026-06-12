"""Tests for the pool standings workbook parser (#20): synthetic workbook."""

import openpyxl
import pytest

from g_nfl.pool.parser import parse_sheet, parse_workbook


def _regular_week_sheet(wb, name="Week 1"):
    """Minimal regular-season sheet: 2 pickers, full slot row, lines block."""
    ws = wb.create_sheet(name)
    headers = ["BB", 2, 3, 4, 5, 6, "U-DOG", "SD", "MON"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=5, column=i, value=h)
    # pick grid
    griffin = ["KC", "DET", "TB", "NE", "BUF", "HOU", "NO", "DEN", "CHI"]
    harry = ["LAC", "GB", "ATL", "NE", "BUF", "HOU", "NO", "--", "MIN"]
    for col, v in enumerate(griffin, start=2):
        ws.cell(row=6, column=col, value=v)
    ws.cell(row=6, column=1, value="Griffin")
    for col, v in enumerate(harry, start=2):
        ws.cell(row=7, column=col, value=v)
    ws.cell(row=7, column=1, value="Harry")
    # result grid: headers mirror pick headers at col 27+
    for i, h in enumerate(headers, start=27):
        ws.cell(row=5, column=i, value=h)
    ws.cell(row=6, column=26, value="Griffin")
    for i, r in enumerate("WLWLWLWLW", start=27):
        ws.cell(row=6, column=i, value=r)
    # lines block: away | line (away team's spread) | home
    ws.cell(row=20, column=2, value="Away")
    lines = [
        ("KC", -3.0, "LAC"),
        ("DET", 2.5, "GB"),
        ("TB", "-1.5", "ATL"),  # line stored as text
        ("NE", -7.0, "BUF"),
        ("HOU", 3.0, "LA"),
        ("NO", 6.5, "DEN"),
        ("MIN", -1.5, "CHI"),
    ]
    for r, (a, line, h) in enumerate(lines, start=21):
        ws.cell(row=r, column=2, value=a)
        ws.cell(row=r, column=3, value=line)
        ws.cell(row=r, column=4, value=h)
    return ws


def _superbowl_sheet(wb):
    ws = wb.create_sheet("SuperBowl")
    ws.cell(row=5, column=3, value="Spread")
    ws.cell(row=5, column=4, value="Total")
    ws.cell(row=6, column=1, value="Griffin")
    ws.cell(row=6, column=3, value="SEA")
    ws.cell(row=6, column=4, value="OVER ")  # trailing space, as in real sheet
    ws.cell(row=20, column=2, value="Away")
    ws.cell(row=21, column=2, value="NE")
    ws.cell(row=21, column=3, value=4.5)
    ws.cell(row=21, column=4, value="SEA")
    ws.cell(row=22, column=2, value="OVER")
    ws.cell(row=22, column=3, value=45.5)
    ws.cell(row=22, column=4, value="UNDER")
    return ws


@pytest.fixture
def workbook(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _regular_week_sheet(wb)
    _superbowl_sheet(wb)
    wb.create_sheet("Standings")  # non-week sheet must be ignored
    path = tmp_path / "standings.xlsx"
    wb.save(path)
    return path


def test_parse_workbook_weeks_and_counts(workbook):
    picks = parse_workbook(workbook, season=2025)
    weeks = {p["week"] for p in picks}
    assert weeks == {1, 22}  # SuperBowl -> week 22, Standings ignored
    wk1 = [p for p in picks if p["week"] == 1]
    # Griffin 9 picks, Harry 8 (survivor '--' skipped)
    assert len([p for p in wk1 if p["picker"] == "Griffin"]) == 9
    assert len([p for p in wk1 if p["picker"] == "Harry"]) == 8


def test_slots_and_pick_types(workbook):
    picks = parse_workbook(workbook, season=2025)
    griffin = {
        p["slot"]: p for p in picks if p["picker"] == "Griffin" and p["week"] == 1
    }
    assert griffin["bb"]["pick_type"] == "best_bet"
    assert griffin["2"]["pick_type"] == "regular"
    assert griffin["udog"]["pick_type"] == "underdog"
    assert griffin["sd"]["pick_type"] == "survivor"
    assert griffin["mnf"]["pick_type"] == "mnf"


def test_spread_sign_and_game_id(workbook):
    picks = parse_workbook(workbook, season=2025)
    griffin = {
        p["slot"]: p for p in picks if p["picker"] == "Griffin" and p["week"] == 1
    }
    # KC away at -3: picked team's spread is the away line
    assert griffin["bb"]["team_picked"] == "KC"
    assert griffin["bb"]["spread"] == -3.0
    assert griffin["bb"]["game_id"] == "2025_01_KC_LAC"
    # CHI home vs MIN -1.5: home perspective flips the sign
    assert griffin["mnf"]["team_picked"] == "CHI"
    assert griffin["mnf"]["spread"] == 1.5
    # text line '-1.5' parsed for TB
    assert griffin["3"]["spread"] == -1.5


def test_results_mapped_from_wl_grid(workbook):
    picks = parse_workbook(workbook, season=2025)
    griffin = {
        p["slot"]: p for p in picks if p["picker"] == "Griffin" and p["week"] == 1
    }
    assert griffin["bb"]["result"] == "W"
    assert griffin["2"]["result"] == "L"
    assert griffin["mnf"]["result"] == "W"
    # Harry has no result row in the grid
    harry = [p for p in picks if p["picker"] == "Harry"]
    assert all(p["result"] is None for p in harry)


def test_superbowl_total_pick(workbook):
    picks = parse_workbook(workbook, season=2025)
    sb = {p["slot"]: p for p in picks if p["week"] == 22}
    assert sb["spread"]["team_picked"] == "SEA"
    assert sb["spread"]["spread"] == -4.5  # SEA home, NE +4.5 away
    assert sb["spread_total"]["team_picked"] == "OVER"
    assert sb["spread_total"]["pick_type"] == "total"
    assert sb["spread_total"]["spread"] == 45.5  # over/under line
    assert sb["spread_total"]["game_id"] == "2025_22_NE_SEA"


def test_team_standardization(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Week 2")
    ws.cell(row=5, column=2, value="BB")
    ws.cell(row=6, column=1, value="Griffin")
    ws.cell(row=6, column=2, value="LAR")  # sheet code -> nflverse 'LA'
    ws.cell(row=20, column=2, value="Away")
    ws.cell(row=21, column=2, value="LAR")
    ws.cell(row=21, column=3, value=-3.0)
    ws.cell(row=21, column=4, value="JAC")  # -> JAX
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    picks = parse_workbook(path, season=2025)
    assert picks[0]["team_picked"] == "LA"
    assert picks[0]["game_id"] == "2025_02_LA_JAX"


def test_unknown_header_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Week 3"
    ws.cell(row=5, column=2, value="WEIRD")
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="unrecognized pick header"):
        parse_workbook(path, season=2025)


def test_parse_sheet_no_lines_block(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=5, column=2, value="BB")
    ws.cell(row=6, column=1, value="Griffin")
    ws.cell(row=6, column=2, value="KC")
    picks = parse_sheet(ws, season=2025, week=1, week_label="Week 1")
    assert picks[0]["game_id"] is None
    assert picks[0]["spread"] is None
